import shutil
import openpyxl
import re
import zipfile
from pathlib import Path
from typing import Dict, Optional


class DemographicsMerger:
    SHEET_MAPPINGS = {
        '1_mile': {
            'Summary': '1 mile summary tab',
            'Demographic Quick Facts': '1 mile demo quick',
            '2025': '1 mile 2025',
            '2030': '1 mile 2030',
            'Income by Age of Household': '1 mile age of HHer'
        },
        '3_mile': {
            'Summary': '3 mile summary',
            'Demographic Quick Facts': '3 mile demo quick',
            '2025': '3 mile 2025',
            '2030': '3 mile 2030',
            'Income by Age of Household': '3 mile age of HHer'
        },
        '5_mile': {
            'Summary': '5 mile summary',
            'Demographic Quick Facts': '5 mile demo quick',
            '2025': '5 mile 2025',
            '2030': '5 mile 2030',
            'Income by Age of Household': '5 mile age of HHer'
        }
    }

    def __init__(self, template_path: str):
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")

    def merge_reports(self, mile_1_path: str, mile_3_path: str, mile_5_path: str,
                      output_path: str, address: Optional[str] = None) -> bool:
        try:
            shutil.copy2(self.template_path, output_path)
            wb_output = openpyxl.load_workbook(output_path)

            files = {
                '1_mile': mile_1_path,
                '3_mile': mile_3_path,
                '5_mile': mile_5_path
            }

            for radius, file_path in files.items():
                if not Path(file_path).exists():
                    continue
                self._copy_sheets_from_file(
                    source_path=file_path,
                    dest_workbook=wb_output,
                    mappings=self.SHEET_MAPPINGS[radius]
                )

            if address and 'Summary' in wb_output.sheetnames:
                wb_output['Summary']['A1'] = address

            wb_output.save(output_path)
            return True

        except Exception as e:
            print(f"ERROR: Failed to merge reports: {e}")
            return False

    def _copy_sheets_from_file(self, source_path: str, dest_workbook, mappings: Dict[str, str]):
        wb_source = openpyxl.load_workbook(source_path, data_only=True)

        for source_name, dest_name in mappings.items():
            if source_name not in wb_source.sheetnames or dest_name not in dest_workbook.sheetnames:
                continue
            self._copy_sheet_data(wb_source[source_name], dest_workbook[dest_name])

        wb_source.close()

    def _copy_sheet_data(self, ws_source, ws_dest):
        ws_dest.delete_rows(1, ws_dest.max_row)
        ws_dest.delete_cols(1, ws_dest.max_column)

        for row in ws_source.iter_rows():
            for cell in row:
                dest_cell = ws_dest.cell(row=cell.row, column=cell.column)
                if cell.value is not None:
                    dest_cell.value = cell.value
                if cell.has_style:
                    dest_cell.font = cell.font.copy()
                    dest_cell.border = cell.border.copy()
                    dest_cell.fill = cell.fill.copy()
                    dest_cell.number_format = cell.number_format
                    dest_cell.protection = cell.protection.copy()
                    dest_cell.alignment = cell.alignment.copy()

        for col in ws_source.column_dimensions:
            ws_dest.column_dimensions[col].width = ws_source.column_dimensions[col].width

        for row in ws_source.row_dimensions:
            ws_dest.row_dimensions[row].height = ws_source.row_dimensions[row].height

        for merged_range in ws_source.merged_cells.ranges:
            ws_dest.merge_cells(str(merged_range))


def validate_file_radius(filename: str) -> Optional[str]:
    name = filename.lower()
    if '1_mi' in name or '1 mi' in name or '1mi' in name:
        return '1'
    elif '3_mi' in name or '3 mi' in name or '3mi' in name:
        return '3'
    elif '5_mi' in name or '5 mi' in name or '5mi' in name:
        return '5'
    return None


def extract_address_from_filename(filename: str) -> Optional[str]:
    name = Path(filename).stem
    patterns = [
        r'Dashboard\s*[-_]\s*(.+?)\s*[-_]\s*[135]\s*mi',
        r'Dashboard\s*[-_]\s*(.+?)\s*[-_]\s*[135]_mi',
        r'[-_]\s*(.+?)\s*[-_]\s*[135]\s*mi',
    ]
    for pattern in patterns:
        match = re.search(pattern, name, re.IGNORECASE)
        if match:
            address = match.group(1).replace('_', ' ')
            address = re.sub(r'\s+', ' ', address).strip(' -_')
            return address
    return None
